import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import csv

epoch = 5
scale = 3

def run_srfbn(filename):
    #运行的python程序的路径
    python_script="train.py"

    # 将 Python 程序路径和参数合并为一个列表
    # command = ["python", python_script, "-opt", f"options/train/x{scale}/{filename}.json", "-param", str(num)]
    command = ["python", python_script, "-opt", f"options/train/RDN_Test/{filename}.json", "-param", str(num)]
    subprocess.run(command, bufsize=0)
    
    #subprocess.run(["python", "train.py", "-opt", f"options/train/{filename}.json", "-param", num],bufsize=0)
    #subprocess.run(["python", "train.py", "-opt", f"options/train/{filename}.json"], bufsize=0)


def opencsv(name, row_index, col_index):
    with open(f"./experiments/Cam{name}_experiments/records/train_records.csv", 'r', encoding='UTF8') as file:
        # 创建 CSV 阅读器对象
        reader = csv.reader(file)

        row_index = row_index - 1
        col_index = col_index - 1

        # 遍历 CSV 文件中的每一行
        for i, row in enumerate(reader):
            # 如果当前行是指定的行
            if i == row_index:
                # 获取指定单元格的值
                csv_content = row[col_index]
                break  # 找到目标单元格后退出循环
        return csv_content


def write_excel(sheet, markrow, cam_num, row1):
    for n in range(1, epoch+1+row1):
        for i in range(1, 7):
            col = i + (cam_num - 1) * 7
            sheet.cell(row=n+markrow-row1, column=col).value = opencsv(cam_num, n+1-row1, i)


def run_avg():
    subprocess.run(["python","avg.py"],bufsize=0)


if __name__=='__main__':
    workbook = Workbook()
    sheet =  workbook.active
    markrow = 1
    print("------======Round : [ 0 ] Start======------")
    for j in range(1,6):
        print("------======Round : [ 0 ] Cam%d ======------" % (j))
        num = j
        filename = f"Cam{j}_start"
        run_srfbn(filename)
        write_excel(sheet, markrow, j, 1)
    markrow = markrow + epoch
    print("------======Round : [ 0 ] Finish======------")
    for i in range(1,10):
        run_avg()
        print("------======Round : [ %d ] Start======------" % (i))
        for j in range(1,2):
            print("------======Round : [ %d ] Cam%d ======------" % (i,j))
            num = j
            filename = f"Cam{j}"
            run_srfbn(filename)
            write_excel(sheet, markrow, j, 0)
        markrow = markrow + epoch
        workbook.save(f"training_record_x{scale}.xlsx")
        print("------======Round : [ %d ] Finish======------" % (i))



